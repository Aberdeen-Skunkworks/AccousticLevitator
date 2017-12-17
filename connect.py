#!/usr/bin/python

class Controller():
    def __init__(self):
        import serial
        import os
        if os.name == "nt":
            self.com = serial.Serial(port="COM3", baudrate=460800, timeout=0.5)
        else:
            self.com = serial.Serial(port="/dev/ttyUSB0", baudrate=460800, timeout=0.5)
            
    def getOutputs(self):
        self.com.write(bytearray([0b11000000,0,0]))
        ack = bytearray(self.com.read(1))[0]
        return ack        
    
    def sendCmd(self, bytestream):
        #Serial communication is carried out using 8 bit/byte
        #chunks. To be able to communicate all the data we need, we
        #use commands composed of 3 bytes. We use the highest bit of
        #each byte to denote if it is the first byte in a command.
        #The remaining bytes must not have the uppermost bit set (to
        #prevent errors in communication causing the controller to
        #assume they are the start of a command), thus only 21bits of
        #information are available for each command.

        # #The structure of a command
        #   23  22  21  20  19  18  17  16  15  14  13  12  11  10  9   8   7   6   5   4   3   2   1   0
        # | 1 | X | X | X | X | X | X | X | 0 | X | X | X | X | X | X | X | 0 | X | X | X | X | X | X | X |
        # 
        self.com.write(bytestream)
        
    def loadOffsets(self):
        self.sendCmd(bytearray([0b10100000, 0, 0]))

    def setOffset(self, clock, offset):
        # #The structure of the command
        #   23  22  21  20  19  18  17  16  15  14  13  12  11  10  9   8   7   6   5   4   3   2   1   0
        # | 1 | 0 | 0 | X | X | X | X | X | 0 | X | X | X | Z | Y | Y | Y | 0 | Y | Y | Y | Y | Y | Y | Y |
        #  X = 8 bit clock select
        #  Y = 10 bit half-phase offset
        #  Z = phase of first oscillation
        sign = 0
        offset = offset % 1250
        if offset > 624:
            sign = 1
            offset = offset - 625
        low_offset =  offset & 0b01111111
        high_offset = ((offset >> 7) & 0b00000111) + (sign << 3)

        b1 = 0b10000000 | (clock>>3)
        b2 = ((clock & 0b00000111)<<4)  | high_offset
        b3 = low_offset
        cmd = bytearray([b1, b2, b3])
        self.sendCmd(cmd)

    def benchmark(self):
        import timeit
        start = timeit.default_timer()
        NTests = 1000
        outputs = self.getOutputs()
        for i in range (NTests):
            for i in range(outputs):
                ctl.setOffset(i, 0)
            ctl.loadOffsets()
        end = timeit.default_timer()
        print( "Benchmark - Pattern update at ", NTests/float(end-start), "Hz")
        
ctl = None
ctl = Controller()
print ("Connected to controller with", ctl.getOutputs(), "outputs.")



def read_from_excel_phase():
    #import required libraries
    from openpyxl import load_workbook; import numpy as np
    #read  from excel file
    wb = load_workbook('phase.xlsx')
    sheet_1 = wb.get_sheet_by_name('phase')
    phase = np.zeros(88)
    for i in range(0,88):
        phase[i]=sheet_1.cell(row=i+1, column=14).value
    return phase


## Testing to see deleted transducers visulised ##
# -------------------------Import Libaries------------------------------------
import numpy as np; import transducer_placment; import math; import phase_algorithms

trans_to_delete = []  # List of unwanted transducers leave blank to keep all
rt = transducer_placment.big_daddy()    # spcing , x nummber, y number of transducers
rt = transducer_placment.delete_transducers(rt,trans_to_delete)

ntrans = len(rt)
x = np.zeros(ntrans)
y = np.zeros(ntrans)
for transducer in range (0,ntrans): # Writing the coordinates to output rt
    x[transducer]= rt[transducer,0,0]
    y[transducer]= rt[transducer,0,2]
"""
phase_index = np.zeros((ntrans),dtype=int)
phi_focus = read_from_excel_phase()
for transducer in range(0,ntrans):
    phase_index[transducer] = int(2500-1250*phi_focus[transducer]/(2*math.pi))
"""
    
outputs = ctl.getOutputs()
t=0
while True:
    phase_index = np.zeros((ntrans),dtype=int)
    t=t+1
    phi_focus = phase_algorithms.phase_find(rt,0,0.05+0.03*math.sin(t*math.pi/100),0)
    for transducer in range(0,ntrans):
        phase_index[transducer] = 2500-int(phi_focus[transducer]/((2*math.pi)/1250))
    for i in range(outputs):
        ctl.setOffset(i,phase_index[i])
    ctl.loadOffsets()
    for i in range(outputs):
        ctl.setOffset(i,0)
    ctl.loadOffsets()

