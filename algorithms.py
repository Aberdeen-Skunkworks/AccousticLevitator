
def get_angle(rt): 
    
# takes in co-odinates of transducers and outputs the angles between 0 and 2 pi 
# angles are anticlockwise from the x axis for each transducer in turn
    
    import numpy as np; import math;
    
    x = rt[:,0]
    y = rt[:,1]
    
    #rho = np.sqrt(x**2 + y**2)
    phi = np.arctan2(y, x)
    
    phi[phi<0] += math.pi*2

    #print(np.divide(phi,np.divide(math.pi,180))) # To test uncomment to get angle in degrees 
    return phi


def pressure (r, rt, phi, nt):
    import numpy as np; import math; import cmath; import constants
    mag_nt = np.linalg.norm(nt)
    nt = np.divide(nt, mag_nt)  # normalising the vector to the unit direction vector

    rs = np.subtract(r,rt)
    k = (2*math.pi)/(float(constants.lamda))   
    mag_rs = np.linalg.norm(rs)
    if mag_rs < 0.001: # Setting calculated dp/dr on top of transducer to zero.
        pressure = 0+0j
    else:
        theta = math.acos((np.dot(rs,nt)) / (mag_rs))
        exponential =  cmath.exp( 1j * (phi + k * mag_rs) ) / mag_rs
        if theta < 0.0000001: # zero angle causes divistion by zero error -> directionality function aproaches 0
            theta = 0.0000001
        frac =  (constants.p0*constants.A)*( (math.sin(k*constants.a*math.sin(theta))) /  (k*constants.a*math.sin(theta)) )
        pressure = (exponential*frac)
    return pressure


def differentiate_pressure(r, rt, phi, nt):
    # Differentiation of pressure with respect to (x,y,z) done analytically
    import math; import cmath; import numpy as np; import constants
    
    mag_nt = np.linalg.norm(nt)
    nt = np.divide(nt, mag_nt)  # normalising the vector to the unit direction vector
    
    rs = np.subtract(r,rt)
    mag_rs = np.linalg.norm(rs)
    if mag_rs < 0.001: # Setting calculated dp/dr on top of transducer to zero.
        d_p_dr = [ 0, 0, 0] 
    else:
        rs_hat = rs / mag_rs
        k = (2*math.pi)/(float(constants.lamda))  
        theta = math.acos((np.dot(rs,nt)) / (mag_rs))
        exponent_term =  cmath.exp( 1j * (phi + k * mag_rs) ) / mag_rs
        if theta < 0.0001: # zero angle causes divistion by zero error -> directionality function aproaches 0
            fraction_term =  (constants.p0*constants.A)
            d_exponential_term_dr = ((rs_hat * 1j * k * cmath.exp(1j*(phi + k * mag_rs)) ) / (mag_rs))  -  ((cmath.exp(1j*(phi + k * mag_rs)) * rs_hat) / (mag_rs**2) )
            d_p_dr = (fraction_term * d_exponential_term_dr)
        else:
            fraction_term =  (constants.p0*constants.A)*( (math.sin(k*constants.a*math.sin(theta))) /  (k*constants.a*math.sin(theta)) )
            d_exponential_term_dr = ((rs_hat * 1j * k * cmath.exp(1j*(phi + k * mag_rs)) ) / (mag_rs))  -  ((cmath.exp(1j*(phi + k * mag_rs)) * rs_hat) / (mag_rs**2) )
            numerator = constants.p0 * constants.A * math.sin(k*constants.a*math.sin(theta))
            denominator =  k*constants.a*math.sin(theta)
            d_denominator_dr = k*constants.a*(( ((-np.dot(rs,nt))/(mag_rs) ) / ((1 - (np.dot(rs,nt))/(mag_rs) )**0.5) ) * ( (np.dot(nt,mag_rs) - (rs_hat*np.dot(nt,rs)) ) / (mag_rs**2) ) )
            d_numerator_dr = constants.p0 * constants.A * math.cos(k*constants.a*math.sin(theta))*d_denominator_dr
            d_fraction_dr = ((d_numerator_dr*denominator) - (numerator*d_denominator_dr))/(denominator**2)
                
            d_p_dr = (d_fraction_dr * exponent_term) + (fraction_term * d_exponential_term_dr)
    return d_p_dr


def acoustic_potential (r, rt, phi, nt):
    
    # Takes in any number of transducer positions and phases and a postion in space
    # will output the acoustic potential U at that location
    
    import numpy as np; import constants; import algorithms
    ntrans = len(rt)
    pressure = np.zeros((ntrans), dtype=complex)
    differentiate_pressure_x = np.zeros((ntrans), dtype=complex)
    differentiate_pressure_y = np.zeros((ntrans), dtype=complex)
    differentiate_pressure_z = np.zeros((ntrans), dtype=complex)
    
    for transducer in range (0, ntrans):
        nt_trans = [ nt[transducer,0], nt[transducer,1], nt[transducer,2] ]
        rt_trans = [ rt[transducer,0], rt[transducer,1], rt[transducer,2] ]
        phi_trans = phi[transducer]
        
        pressure[transducer] = algorithms.pressure(r, rt_trans , phi_trans, nt_trans)
        differentiate_pressure = algorithms.differentiate_pressure(r, rt_trans , phi_trans, nt_trans)
        differentiate_pressure_x[transducer] = differentiate_pressure[0]
        differentiate_pressure_y[transducer] = differentiate_pressure[1]
        differentiate_pressure_z[transducer] = differentiate_pressure[2]
    
    total_pressure = np.sum(pressure)
    total_diff_pressure_x = np.sum(differentiate_pressure_x)
    total_diff_pressure_y = np.sum(differentiate_pressure_y)
    total_diff_pressure_z = np.sum(differentiate_pressure_z)

    p_abs   = np.absolute(total_pressure)  
    px_abs  = np.absolute(total_diff_pressure_x)  
    py_abs  = np.absolute(total_diff_pressure_y)  
    pz_abs  = np.absolute(total_diff_pressure_z)  

    left_side = np.multiply(np.power(p_abs, 2), constants.m1)
    gradients = np.add(np.add(np.power(px_abs, 2), np.power(py_abs, 2)), np.power(pz_abs, 2))
    right_side =  np.multiply(constants.m2, gradients)
    u = np.subtract(left_side, right_side)

    #### add this back in when you need absolute pressure , p_abs
    return u


def differentiate_acoustic_potential(h,r,rt,phi,nt):
    x = r[0]; y = r[1]; z = r[2]
    
    du_dr_numercal_x = (acoustic_potential( [x+h, y, z], rt,phi,nt ) - acoustic_potential( [x-h, y, z], rt,phi,nt)) / (2*h)
    du_dr_numercal_y = (acoustic_potential( [x, y+h, z], rt,phi,nt ) - acoustic_potential( [x, y-h, z], rt,phi,nt)) / (2*h)
    du_dr_numercal_z = (acoustic_potential( [x, y, z+h], rt,phi,nt ) - acoustic_potential( [x, y, z-h], rt,phi,nt)) / (2*h)
    du_dr_numercal   = [du_dr_numercal_x, du_dr_numercal_y, du_dr_numercal_z]

    return du_dr_numercal


def read_from_excel(): # Reads the transducer locations from excel locations.xlsx
    #import required libraries
    from openpyxl import load_workbook
    import numpy as np
    #read  from excel file
    wb = load_workbook('locations.xlsx')
    sheet_1 = wb.get_sheet_by_name('Sheet1')
    x = np.zeros(88)
    y = np.zeros(88)
    coordinates = np.zeros((88,2))
    for i in range(0,88):
        x[i]=sheet_1.cell(row=i+9, column=10).value
        y[i]=sheet_1.cell(row=i+9, column=11).value
     
    coordinates = [x,y]
    return coordinates
"""
import matplotlib.pyplot as plt;
test = read_from_excel()
plt.plot(test[0], test[1], 'ro')
plt.show()
"""



def circle_co_ords(splits, diameter): 
    """Take in number of points and circle diamiter and output coordinates of points in a circle centered at the origin"""
    import math; import numpy as np
    x = np.zeros (splits)
    y = np.zeros (splits)
    coordinates = np.zeros((splits,2))
    for point in range(0,splits):
        angle = ((2*math.pi) / splits) * (point+1)
        x[point] = diameter * math.cos(angle)
        y[point] = diameter * math.sin(angle)
    coordinates = [x,y]
    return coordinates
"""
import matplotlib.pyplot as plt;
test = circle_co_ords(50, 0.01)
plt.plot(test[0], test[1], 'ro')
plt.show()
"""

def read_from_excel_phases(): # Reads the phases from phase.xlsx
    #import required libraries
    from openpyxl import load_workbook; import numpy as np
    #read  from excel file
    wb = load_workbook('phase.xlsx')
    sheet_1 = wb.get_sheet_by_name('phase')
    phases = np.zeros((88))
    for i in range(0,88):
        phases[i]=sheet_1.cell(row=i+1, column=14).value        
    return phases




def force_calc(focus_point, rt, nt, phi, vti = False): # Set vti to True to output vti files
    
    import constants; import numpy as np; import calc_pressure_field; import time;
    from vti_writer import vti_writer; import scipy.ndimage
    
    z_distances = np.linspace(-constants.gsize + focus_point[2],   constants.gsize + focus_point[2], constants.npoints)
    
    
    # ----------------------Setting up output arrays-------------------------------
    
    pcombined = np.zeros ((constants.npoints,constants.npoints,constants.npoints),dtype=complex)
    
    px = np.zeros ((constants.npoints,constants.npoints,constants.npoints), dtype=complex)
    py = np.copy(px); pz = np.copy(px)
    
    ux = np.zeros ((constants.npoints,constants.npoints,constants.npoints), dtype=float)
    uy = np.copy(ux); uz = np.copy(ux)
    
    pabs = np.zeros ((constants.npoints,constants.npoints,constants.npoints), dtype=float) 
    pxabs = np.copy(pabs); pyabs = np.copy(pabs); pzabs = np.copy(pabs)
    
    u = np.zeros ((constants.npoints,constants.npoints,constants.npoints), dtype=float)
    
    height = np.zeros ((constants.npoints,constants.npoints,constants.npoints), dtype=float)
    
    ntrans = len(rt)
    # ----------------------------------------------------------------------------
    

    t1 = time.time()
    p = calc_pressure_field.calc_pressure_field_numpy_around_trap(rt, nt, ntrans, phi, focus_point)
    t2 = time.time()
    print(" ")
    print("Numpy pressure calculation took ",round(t2-t1,2), " seconds" )
    print(" ")
    
    
    # -----------------Loop to sum pressure of all transducers---------------------
    
    height_2D = np.tile(z_distances, (constants.npoints ,1))
    height = np.broadcast_to(height_2D, (constants.npoints, constants.npoints, constants.npoints))
    pcombined = np.sum(np.copy(p), axis = 0) 
    
    # -----------------Calculating derrivitive of the pressure field---------------
    
    diff_p = np.gradient(pcombined,constants.deltaxyz)
    px = np.copy(diff_p[0]); py = np.copy(diff_p[1]); pz = np.copy(diff_p[2])
    
    # ----------------- Calculating the Gork’ov potential-------------------------
    
    # Calculating absuloute values for the pressure field and its derrivitives
           
    pabs   = np.absolute(pcombined)  
    pxabs  = np.absolute(px)  
    pyabs  = np.absolute(py)  
    pzabs  = np.absolute(pz)  
    
    # Calculating Gork’ov potential u (Including gravity)
    """ ################### WRONG
    u = ( np.subtract( np.subtract( np.multiply(np.multiply(2, constants.k1), np.power(pabs, 2) ), 
    np.multiply( np.multiply(2, constants.k2), np.add(np.add(np.power(pxabs, 2), np.power(pyabs, 2)), np.power(pzabs, 2) ) ) ),  
    np.multiply(constants.p_mass, np.multiply(constants.gravity, np.copy(height)))) ) # Gravity term
    """
    
    left_side = np.multiply(np.power(pabs, 2), constants.m1)
    gradients = np.add(np.add(np.power(pxabs, 2), np.power(pyabs, 2)), np.power(pzabs, 2))
    right_side =  np.multiply(constants.m2, gradients)
    u = np.subtract(left_side, right_side)
    
    u_with_gravity = np.subtract(u, np.multiply(constants.p_mass, np.multiply(constants.gravity, np.copy(height))))
    u_with_gravity_nano = np.multiply(u_with_gravity, 1000000000)

    # -----------------Calculating derrivitive of Gork’ov potential ---------------
    
    diff_u = np.gradient(u_with_gravity,constants.deltaxyz)
    ux = np.copy(diff_u[0]); uy = np.copy(diff_u[1]); uz = np.copy(diff_u[2])
    
    # -----------------Calculating force on particle ---------------
    
    
    fx = np.copy(-ux); fy = np.copy(-uy); fz = np.copy(-uz)
    
    
    laplace_u = scipy.ndimage.filters.laplace(u_with_gravity)
    

    # -------------------Creating output images and vtk file----------------------
    
    if vti == True:
        vti_writer(constants.npoints, pabs, fx, fy, fz, u_with_gravity_nano, laplace_u)
    
    return pabs, fx, fy, fz, u_with_gravity, u_with_gravity_nano, laplace_u